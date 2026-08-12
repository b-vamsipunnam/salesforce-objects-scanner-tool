import json
import subprocess
import sys
import tempfile
import unittest
import xml.etree.ElementTree as ET
import zipfile
from pathlib import Path

from openpyxl import load_workbook
from src.robot.libraries.SfUtils import get_sf_error_details


class ExcelWriterTests(unittest.TestCase):
    def test_generates_formatted_workbook(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            inputs = [
                {"Account": 100},
                {"ApexClass": 4},
                {"AggregateResult": "INVALID_TYPE"},
                {
                    "AggregateResult": {
                        "name": "INVALID_TYPE",
                        "message": "sObject type is not supported",
                    }
                },
                {"Account": 0.25},
            ]
            paths = []
            for index, payload in enumerate(inputs):
                path = root / f"{index}.json"
                path.write_text(json.dumps(payload), encoding="utf-8")
                paths.append(path)
            output = root / "report.xlsx"
            script = Path("src/robot/libraries/ExcelWriter.py")
            result = subprocess.run(
                [sys.executable, str(script), *(str(path) for path in paths), str(output)],
                capture_output=True,
                text=True,
                check=False,
            )
            self.assertEqual(result.returncode, 0, result.stderr)
            workbook = load_workbook(output)
            self.assertEqual(
                workbook.sheetnames,
                ["Data Objects", "Tooling Objects", "Skipped Objects", "Durations (Seconds)"],
            )
            for sheet in workbook.worksheets:
                self.assertEqual(sheet.freeze_panes, "A2")
                self.assertIsNone(sheet.auto_filter.ref)
                self.assertTrue(sheet.tables)
            self.assertEqual(workbook["Data Objects"]["B2"].value, 100)
            self.assertIn("INVALID_TYPE", workbook["Skipped Objects"]["C2"].value)

            namespace = {"x": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
            with zipfile.ZipFile(output) as archive:
                worksheet_names = [
                    name
                    for name in archive.namelist()
                    if name.startswith("xl/worksheets/sheet") and name.endswith(".xml")
                ]
                table_names = [
                    name for name in archive.namelist() if name.startswith("xl/tables/table")
                ]
                self.assertEqual(len(worksheet_names), 4)
                self.assertEqual(len(table_names), 4)
                for name in worksheet_names:
                    worksheet = ET.fromstring(archive.read(name))
                    self.assertIsNone(worksheet.find("x:autoFilter", namespace))
                for name in table_names:
                    table = ET.fromstring(archive.read(name))
                    self.assertIsNotNone(table.find("x:autoFilter", namespace))

    def test_rejects_non_dictionary_input(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            inputs = [[], {}, {}, {}, {}]
            paths = []
            for index, payload in enumerate(inputs):
                path = root / f"{index}.json"
                path.write_text(json.dumps(payload), encoding="utf-8")
                paths.append(path)
            output = root / "report.xlsx"
            script = Path("src/robot/libraries/ExcelWriter.py")
            result = subprocess.run(
                [sys.executable, str(script), *(str(path) for path in paths), str(output)],
                capture_output=True,
                text=True,
                check=False,
            )
            self.assertNotEqual(result.returncode, 0)
            self.assertFalse(output.exists())

    def test_persisted_json_and_excel_contain_only_redacted_error_details(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            secrets = ("rawBearerSecret", "rawSessionSecret", "rawRefreshSecret")
            raw_error = json.dumps(
                {
                    "name": "EXTERNAL_OBJECT_EXCEPTION",
                    "message": (
                        "Authorization: bearer rawBearerSecret "
                        "sessionId=rawSessionSecret "
                        "force://client:secret:rawRefreshSecret@example.my.salesforce.com"
                    ),
                }
            )
            sanitized = get_sf_error_details(raw_error)
            payloads = [
                {},
                {},
                {"ExternalObject": "EXTERNAL_OBJECT_EXCEPTION"},
                {"ExternalObject": sanitized},
                {"ExternalObject": 1.5},
            ]
            paths = []
            for index, payload in enumerate(payloads):
                path = root / f"persisted-{index}.json"
                path.write_text(json.dumps(payload), encoding="utf-8")
                paths.append(path)

            persisted_details = paths[3].read_text(encoding="utf-8")
            for secret in secrets:
                self.assertNotIn(secret, persisted_details)
            self.assertIn("[REDACTED_BEARER_TOKEN]", persisted_details)
            self.assertIn("[REDACTED_SESSION_ID]", persisted_details)
            self.assertIn("[REDACTED_AUTH_URL]", persisted_details)

            output = root / "redacted-report.xlsx"
            script = Path("src/robot/libraries/ExcelWriter.py")
            result = subprocess.run(
                [sys.executable, str(script), *(str(path) for path in paths), str(output)],
                capture_output=True,
                text=True,
                check=False,
            )
            self.assertEqual(result.returncode, 0, result.stderr)
            error_cell = load_workbook(output)["Skipped Objects"]["C2"].value
            for secret in secrets:
                self.assertNotIn(secret, error_cell)
            self.assertIn("[REDACTED_BEARER_TOKEN]", error_cell)
            self.assertIn("[REDACTED_SESSION_ID]", error_cell)
            self.assertIn("[REDACTED_AUTH_URL]", error_cell)


if __name__ == "__main__":
    unittest.main()
