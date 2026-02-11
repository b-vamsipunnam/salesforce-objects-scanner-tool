# -*- coding: utf-8 -*-

from io import BytesIO
from typing import Any, Dict, Iterator, List, Optional, Tuple
import os

import openpyxl
from openpyxl.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet


class SuchIdIsExistException(Exception):
    """Raised when the document with the identifier is already in the cache."""
    pass


class NoSuchIdException(Exception):
    """Raised when accessing an absent document identifier."""
    pass


class NoOpenedDocumentsException(Exception):
    """Raised in the absence of open documents."""
    pass


class ExcelLibrary(object):
    """Library for working with Excel documents.

    This is a compatibility wrapper intended to be used as a drop-in Robot Framework library.

    IMPORTANT BEHAVIOR (for backward-compatibility with your Robot suite):
    - 'Create Excel Document' argument is treated as a FILE PATH (not a doc_id).
      It creates the workbook and binds the current document id to that path.
    - 'Save Excel Document' can be called with filename OR with no args; it will save
      to the current document id if it is a path.
    """

    ROBOT_LIBRARY_SCOPE = "GLOBAL"

    def __init__(self) -> None:
        self._cache: Dict[str, openpyxl.Workbook] = {}
        self._current_id: Optional[str] = None

    # -------------------------
    # Helper utilities
    # -------------------------

    def _ensure_parent_dir(self, file_path: str) -> None:
        """Create parent directory if file_path includes a directory."""
        file_path = str(file_path)
        dir_path = os.path.dirname(file_path)
        if dir_path:
            os.makedirs(dir_path, exist_ok=True)

    def _get_current_workbook(self) -> openpyxl.Workbook:
        """Checks opened document and returns current workbook."""
        if not self._cache or not self._current_id:
            raise NoOpenedDocumentsException("No opened documents in cache.")
        return self._cache[self._current_id]

    def get_sheet(self, sheet_name: str = None) -> Worksheet:
        """Returns a sheet from the current document."""
        workbook = self._get_current_workbook()
        if sheet_name is None:
            return workbook.active
        sheet_name = str(sheet_name)
        return workbook[sheet_name]

    # -------------------------
    # Core keywords
    # -------------------------

    def create_excel_document(self, doc_id: str) -> str:
        """
        Creates new excel document.

        COMPAT MODE:
        - 'doc_id' is treated as a FILE PATH (your suite passes ${CV_File_Name} etc.).
        - Creates workbook on disk immediately.
        - Uses the file path as the cache key and current id.
        """
        file_path = str(doc_id)

        # If you call Create twice with same path, keep the old behavior: raise.
        if file_path in self._cache:
            raise SuchIdIsExistException(f"Document with such id {file_path} is created.")

        self._ensure_parent_dir(file_path)

        workbook = openpyxl.Workbook()
        # Save immediately so subsequent steps expecting file existence won't break
        workbook.save(filename=file_path)

        self._cache[file_path] = workbook
        self._current_id = file_path
        return self._current_id

    def open_excel_document(self, filename: str, doc_id: str = None) -> str:
        """
        Opens xlsx document file.

        Supports BOTH signatures:
        - Open Excel Document | filename=file.xlsx | doc_id=myid |
        - Open Excel Document | filename=file.xlsx |   (doc_id omitted)

        If doc_id is omitted, uses filename as id.
        """
        filename = str(filename)
        use_id = str(doc_id) if doc_id is not None else filename

        if use_id in self._cache:
            raise SuchIdIsExistException(f"Document with such id {use_id} is opened.")

        workbook = openpyxl.load_workbook(filename=filename)
        self._cache[use_id] = workbook
        self._current_id = use_id
        return self._current_id

    def open_excel_document_from_stream(self, stream: bytes, doc_id: str) -> str:
        """Opens xlsx document from stream."""
        doc_id = str(doc_id)
        if doc_id in self._cache:
            raise SuchIdIsExistException(f"Document with such id {doc_id} is opened.")
        workbook = openpyxl.load_workbook(filename=BytesIO(stream))
        self._cache[doc_id] = workbook
        self._current_id = doc_id
        return self._current_id

    def switch_current_excel_document(self, doc_id: str) -> Optional[str]:
        """Switches current excel document."""
        doc_id = str(doc_id)
        if doc_id not in self._cache:
            raise NoSuchIdException(f"Document with such id {doc_id} is not opened yet.")
        old_name = self._current_id
        self._current_id = doc_id
        return old_name

    def close_current_excel_document(self) -> Optional[str]:
        """Closes current document (removes from cache)."""
        if self._current_id is not None:
            self._cache.pop(self._current_id, None)
            self._current_id = None
        if self._cache:
            self._current_id = list(self._cache.keys())[0]
        return self._current_id

    def close_all_excel_documents(self) -> None:
        """Closes all opened documents."""
        self._cache = {}
        self._current_id = None

    def save_excel_document(self, filename: str = None) -> None:
        """
        Saves the current document to disk.

        COMPAT MODE:
        - If filename is omitted, saves to current_id when it looks like a path.
        - If filename provided, ensures its parent directory exists.
        """
        workbook = self._get_current_workbook()

        target = filename if filename else self._current_id
        if not target:
            raise NoOpenedDocumentsException("No current document to save.")

        target = str(target)
        self._ensure_parent_dir(target)

        workbook.save(filename=target)

        # If user saved to a new filename, rebind id to that file path
        if filename and self._current_id != target:
            # Move cache entry key to new id
            self._cache[target] = workbook
            if self._current_id in self._cache:
                self._cache.pop(self._current_id, None)
            self._current_id = target

    def get_list_sheet_names(self) -> List[str]:
        """Returns a list of sheet names in the current document."""
        workbook = self._get_current_workbook()
        return workbook.sheetnames

    def make_list_from_excel_sheet(self, sheet: Worksheet) -> list:
        """Making list from Excel sheet."""
        data = []
        for row in sheet.values:
            data.append(row)
        return data

    # -------------------------
    # Read keywords
    # -------------------------

    def read_excel_cell(self, row_num: int, col_num: int, sheet_name: str = None) -> Any:
        """Returns content of a cell."""
        row_num = int(row_num)
        col_num = int(col_num)
        sheet = self.get_sheet(sheet_name)
        cell: Cell = sheet.cell(row=row_num, column=col_num)
        return cell.value

    def read_excel_row(
        self, row_num: int, col_offset: int = 0, max_num: int = 0, sheet_name: str = None
    ) -> List[Any]:
        """Returns content of a row from the current sheet of the document."""
        row_num = int(row_num)
        col_offset = int(col_offset)
        max_num = int(max_num)
        sheet = self.get_sheet(sheet_name)

        if max_num <= 0:
            # If not provided, read until last column with values (best-effort)
            max_num = sheet.max_column - col_offset

        row_iter: Iterator[Tuple[Cell]] = sheet.iter_rows(
            min_row=row_num,
            max_row=row_num,
            min_col=1 + col_offset,
            max_col=col_offset + max_num,
        )
        row: Tuple[Cell, ...] = next(row_iter)
        return [cell.value for cell in row]

    def read_excel_column(
        self, col_num: int, row_offset: int = 0, max_num: int = 0, sheet_name: str = None
    ) -> List[Any]:
        """Returns content of a column from the current sheet of the document."""
        col_num = int(col_num)
        row_offset = int(row_offset)
        max_num = int(max_num)
        sheet = self.get_sheet(sheet_name)

        if max_num <= 0:
            max_num = sheet.max_row - row_offset

        row_iter: Iterator[Tuple[Cell, ...]] = sheet.iter_rows(
            min_col=col_num,
            max_col=col_num,
            min_row=1 + row_offset,
            max_row=row_offset + max_num,
        )
        return [row[0].value for row in row_iter]

    # -------------------------
    # Write keywords
    # -------------------------

    def write_excel_cell(self, row_num: int, col_num: int, value: Any, sheet_name: str = None) -> None:
        """Writes value to the cell."""
        row_num = int(row_num)
        col_num = int(col_num)
        sheet = self.get_sheet(sheet_name)
        sheet.cell(row=row_num, column=col_num, value=value)

    def write_excel_row(
        self, row_num: int, row_data: List[Any], col_offset: int = 0, sheet_name: str = None
    ) -> None:
        """Writes a row to the document."""
        row_num = int(row_num)
        col_offset = int(col_offset)
        sheet = self.get_sheet(sheet_name)
        for col_num in range(len(row_data)):
            sheet.cell(row=row_num, column=col_num + col_offset + 1, value=row_data[col_num])

    def write_excel_rows(
        self, rows_data: List[List[Any]], rows_offset: int = 0, col_offset: int = 0, sheet_name: str = None
    ) -> None:
        """Writes a list of rows to the document."""
        for row_num, row_data in enumerate(rows_data):
            self.write_excel_row(row_num + int(rows_offset) + 1, row_data, col_offset, sheet_name)

    def write_excel_column(
        self, col_num: int, col_data: List[Any], row_offset: int = 0, sheet_name: str = None
    ) -> None:
        """Writes the data to a column."""
        col_num = int(col_num)
        row_offset = int(row_offset)
        sheet = self.get_sheet(sheet_name)
        for row_num in range(len(col_data)):
            sheet.cell(column=col_num, row=row_num + row_offset + 1, value=col_data[row_num])
