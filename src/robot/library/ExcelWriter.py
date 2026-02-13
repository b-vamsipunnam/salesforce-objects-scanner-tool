import sys
import json
from openpyxl import Workbook
from openpyxl.styles import Font


def write_headers(sheet, headers):
    for col, header in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=col)
        cell.value = header
        cell.font = Font(bold=True)


def autosize(sheet):
    for column in sheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        sheet.column_dimensions[column_letter].width = max_length + 4


def main():
    if len(sys.argv) != 6:
        print("Usage: excel_writer.py data_json tooling_json skipped_json durations_json output_file")
        sys.exit(1)

    data_results = json.loads(sys.argv[1])
    tooling_results = json.loads(sys.argv[2])
    skipped_reasons = json.loads(sys.argv[3])
    durations_seconds = json.loads(sys.argv[4])
    output_file = sys.argv[5]

    wb = Workbook()

    data_sheet = wb.active
    data_sheet.title = "Data Objects"
    tooling_sheet = wb.create_sheet("Tooling Objects")
    skipped_sheet = wb.create_sheet("Skipped Objects")
    duration_sheet = wb.create_sheet("Durations (Seconds)")

    write_headers(data_sheet, ["Object Name", "Record Count"])
    write_headers(tooling_sheet, ["Tooling Object Name", "Record Count"])
    write_headers(skipped_sheet, ["Object Name", "Skip Reason"])
    write_headers(duration_sheet, ["Object Name", "Duration (Seconds)"])

    for row, (k, v) in enumerate(data_results.items(), 2):
        data_sheet.cell(row=row, column=1).value = k
        data_sheet.cell(row=row, column=2).value = v

    for row, (k, v) in enumerate(tooling_results.items(), 2):
        tooling_sheet.cell(row=row, column=1).value = k
        tooling_sheet.cell(row=row, column=2).value = v

    for row, (k, v) in enumerate(skipped_reasons.items(), 2):
        skipped_sheet.cell(row=row, column=1).value = k
        skipped_sheet.cell(row=row, column=2).value = v

    for row, (k, v) in enumerate(durations_seconds.items(), 2):
        duration_sheet.cell(row=row, column=1).value = k
        duration_sheet.cell(row=row, column=2).value = v

    for sheet in wb.worksheets:
        autosize(sheet)

    wb.save(output_file)
    print(f"Excel report generated: {output_file}")


if __name__ == "__main__":
    main()
